"""
金融数据初始化脚本：
1. 创建 MySQL 数据库 `financial_report_db`，建立复杂关联数据表并插入测试数据
2. 在 `data/test` 目录生成 Excel 测试模板，用于后续填空功能测试
"""
import pymysql
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from utils import ensure_dir, get_logger
from utils.exception_utils import FileOperateException

# ---------------------- 全局配置（与项目现有配置对齐）----------------------
MYSQL_CONFIG = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': 'David7668',
    'charset': 'utf8mb4'
}
DB_NAME = 'financial_report_db'  # 数据库名称（金融报表专用）
EXCEL_TEMPLATE_PATH = './data/test/financial_cost_fill_template.xlsx'  # Excel模板路径

# ---------------------- 数据库初始化核心逻辑 ----------------------
def init_mysql_database():
    """
    创建数据库 `financial_report_db`，建立3张关联数据表：
    1. department（部门表）：存储企业部门信息（主表）
    2. cost_detail（成本明细表）：存储日常成本数据（关联部门表）
    3. sales_detail（销售明细表）：存储日常销售数据（关联部门表）
    """
    logger = get_logger("MySQL_Init")
    conn = None
    cursor = None

    try:
        # 步骤1：连接MySQL服务（不指定具体数据库）
        conn = pymysql.connect(
            host=MYSQL_CONFIG['host'],
            port=MYSQL_CONFIG['port'],
            user=MYSQL_CONFIG['user'],
            password=MYSQL_CONFIG['password'],
            charset=MYSQL_CONFIG['charset']
        )
        cursor = conn.cursor()
        logger.info("成功连接MySQL服务，开始创建数据库...")

        # 步骤2：创建数据库（如果不存在）
        cursor.execute(f"DROP DATABASE IF EXISTS {DB_NAME};")
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME} DEFAULT CHARACTER SET {MYSQL_CONFIG['charset']};")
        cursor.execute(f"USE {DB_NAME};")
        logger.info(f"数据库 `{DB_NAME}` 创建并切换成功")

        # 步骤3：创建部门表（department）- 主表
        create_department_sql = """
        CREATE TABLE IF NOT EXISTS department (
            dept_id INT PRIMARY KEY AUTO_INCREMENT COMMENT '部门ID（主键）',
            dept_name VARCHAR(50) NOT NULL COMMENT '部门名称',
            dept_level VARCHAR(20) NOT NULL COMMENT '部门级别（如：一级、二级）',
            parent_dept_id INT DEFAULT 0 COMMENT '上级部门ID（0表示无上级）',
            manager VARCHAR(30) COMMENT '部门经理',
            create_time DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
            update_time DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
            is_valid TINYINT DEFAULT 1 COMMENT '是否有效（1：有效，0：无效）',
            INDEX idx_dept_name (dept_name) COMMENT '部门名称索引'
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT '企业部门信息表';
        """
        cursor.execute(create_department_sql)
        logger.info("数据表 `department`（部门表）创建成功")

        # 步骤4：创建成本明细表（cost_detail）- 关联部门表
        create_cost_sql = """
        CREATE TABLE IF NOT EXISTS cost_detail (
            cost_id BIGINT PRIMARY KEY AUTO_INCREMENT COMMENT '成本ID（主键）',
            dept_id INT NOT NULL COMMENT '部门ID（关联department.dept_id）',
            cost_type VARCHAR(50) NOT NULL COMMENT '成本类型（如：人力、物料、运营、房租）',
            cost_amount DECIMAL(16, 2) NOT NULL COMMENT '成本金额（保留2位小数）',
            settle_date DATE NOT NULL COMMENT '结算日期',
            project_name VARCHAR(100) COMMENT '关联项目名称',
            invoice_no VARCHAR(50) COMMENT '发票编号',
            payer VARCHAR(50) COMMENT '付款人',
            remark TEXT COMMENT '备注',
            create_time DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
            update_time DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
            is_valid TINYINT DEFAULT 1 COMMENT '是否有效（1：有效，0：无效）',
            INDEX idx_dept_id (dept_id) COMMENT '部门ID索引',
            INDEX idx_settle_date (settle_date) COMMENT '结算日期索引',
            INDEX idx_cost_type (cost_type) COMMENT '成本类型索引',
            FOREIGN KEY (dept_id) REFERENCES department(dept_id) ON DELETE RESTRICT ON UPDATE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT '企业成本明细表';
        """
        cursor.execute(create_cost_sql)
        logger.info("数据表 `cost_detail`（成本明细表）创建成功")

        # 步骤5：创建销售明细表（sales_detail）- 关联部门表
        create_sales_sql = """
        CREATE TABLE IF NOT EXISTS sales_detail (
            sales_id BIGINT PRIMARY KEY AUTO_INCREMENT COMMENT '销售ID（主键）',
            dept_id INT NOT NULL COMMENT '部门ID（关联department.dept_id）',
            product_type VARCHAR(50) NOT NULL COMMENT '产品类型（如：硬件、软件、服务）',
            sales_amount DECIMAL(16, 2) NOT NULL COMMENT '销售金额（保留2位小数）',
            sale_date DATE NOT NULL COMMENT '销售日期',
            customer_name VARCHAR(100) NOT NULL COMMENT '客户名称',
            region VARCHAR(50) COMMENT '客户区域（如：华北、华东、华南）',
            contract_no VARCHAR(50) COMMENT '合同编号',
            receivable_status VARCHAR(20) COMMENT '回款状态（如：未回款、部分回款、全额回款）',
            remark TEXT COMMENT '备注',
            create_time DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
            update_time DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
            is_valid TINYINT DEFAULT 1 COMMENT '是否有效（1：有效，0：无效）',
            INDEX idx_dept_id (dept_id) COMMENT '部门ID索引',
            INDEX idx_sale_date (sale_date) COMMENT '销售日期索引',
            INDEX idx_customer_name (customer_name) COMMENT '客户名称索引',
            FOREIGN KEY (dept_id) REFERENCES department(dept_id) ON DELETE RESTRICT ON UPDATE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT '企业销售明细表';
        """
        cursor.execute(create_sales_sql)
        logger.info("数据表 `sales_detail`（销售明细表）创建成功")

        # 步骤6：插入测试数据（部门表）
        dept_data = [
            ("总经办", "一级", 0, "张三", 1),
            ("财务部", "一级", 0, "李四", 1),
            ("销售部", "一级", 0, "王五", 1),
            ("技术部", "一级", 0, "赵六", 1),
            ("销售一部", "二级", 3, "钱七", 1),
            ("销售二部", "二级", 3, "孙八", 1),
            ("研发部", "二级", 4, "周九", 1),
            ("运维部", "二级", 4, "吴十", 1)
        ]
        insert_dept_sql = """
        INSERT INTO department (dept_name, dept_level, parent_dept_id, manager, is_valid)
        VALUES (%s, %s, %s, %s, %s);
        """
        cursor.executemany(insert_dept_sql, dept_data)
        logger.info(f"部门表插入 {cursor.rowcount} 条测试数据")

        # 步骤7：插入测试数据（成本明细表，生成30天批量数据）
        cost_data = []
        cost_types = ["人力成本", "物料成本", "运营成本", "房租成本", "水电成本", "办公成本"]
        dept_ids = [1, 2, 3, 4, 5, 6, 7, 8]
        start_date = datetime.now() - timedelta(days=30)

        for i in range(30):
            settle_date = (start_date + timedelta(days=i)).date()
            for dept_id in dept_ids[:4]:  # 前4个部门生成成本数据
                for cost_type in cost_types[:3]:  # 前3种成本类型
                    cost_amount = round((dept_id * 1000) + (i * 100) + (cost_types.index(cost_type) * 500), 2)
                    project_name = f"项目{dept_id}-{i}"
                    invoice_no = f"INV-{dept_id}-{i}-{cost_types.index(cost_type)}"
                    payer = f"付款人{dept_id}"
                    cost_data.append((
                        dept_id, cost_type, cost_amount, settle_date,
                        project_name, invoice_no, payer, f"{cost_type}结算", 1
                    ))

        insert_cost_sql = """
        INSERT INTO cost_detail (dept_id, cost_type, cost_amount, settle_date,
        project_name, invoice_no, payer, remark, is_valid)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        cursor.executemany(insert_cost_sql, cost_data)
        logger.info(f"成本明细表插入 {cursor.rowcount} 条测试数据")

        # 步骤8：插入测试数据（销售明细表，生成30天批量数据）
        sales_data = []
        product_types = ["硬件产品", "软件产品", "技术服务", "咨询服务"]
        regions = ["华北", "华东", "华南", "西北", "西南"]
        receivable_status = ["未回款", "部分回款", "全额回款"]
        dept_ids = [3, 5, 6]  # 销售相关部门

        for i in range(30):
            sale_date = (start_date + timedelta(days=i)).date()
            for dept_id in dept_ids:
                for product_type in product_types[:2]:  # 前2种产品类型
                    sales_amount = round((dept_id * 5000) + (i * 200) + (product_types.index(product_type) * 1000), 2)
                    customer_name = f"客户{dept_id}-{i}"
                    region = regions[i % len(regions)]
                    contract_no = f"CON-{dept_id}-{i}-{product_types.index(product_type)}"
                    receivable = receivable_status[i % len(receivable_status)]
                    sales_data.append((
                        dept_id, product_type, sales_amount, sale_date,
                        customer_name, region, contract_no, receivable, f"{product_type}销售", 1
                    ))

        insert_sales_sql = """
        INSERT INTO sales_detail (dept_id, product_type, sales_amount, sale_date,
        customer_name, region, contract_no, receivable_status, remark, is_valid)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        cursor.executemany(insert_sales_sql, sales_data)
        logger.info(f"销售明细表插入 {cursor.rowcount} 条测试数据")

        # 步骤9：提交事务
        conn.commit()
        logger.info("所有数据库操作提交成功，金融数据初始化完成！")

    except pymysql.Error as e:
        if conn:
            conn.rollback()
        raise Exception(f"MySQL 操作失败：{e.args[0]} - {e.args[1]}")
    finally:
        # 关闭连接
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        logger.info("MySQL 连接已关闭")

# ---------------------- Excel模板生成核心逻辑 ----------------------
def generate_excel_test_template():
    """
    生成Excel测试模板（`data/test/financial_cost_fill_template.xlsx`）
    模板结构：成本汇总表，预留填空单元格（用于后续Office Parser自动填空）
    """
    logger = get_logger("Excel_Template_Init")

    try:
        # 步骤1：确保目录存在
        ensure_dir(os.path.dirname(EXCEL_TEMPLATE_PATH))
        logger.info(f"Excel模板目录已就绪，路径：{os.path.dirname(EXCEL_TEMPLATE_PATH)}")

        # 步骤2：创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "2024年月度成本汇总表"

        # 步骤3：设置模板样式（标题、表头、填空单元格标注）
        # 标题样式
        title_font = Font(name="微软雅黑", size=16, bold=True)
        title_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        ws.merge_cells('A1:H1')
        ws['A1'] = "2024年月度成本汇总表（自动填空测试模板）"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A1'].fill = title_fill

        # 表头样式
        header_font = Font(name="微软雅黑", size=12, bold=True)
        header_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        headers = [
            "部门名称", "人力成本均值", "物料成本均值", "运营成本均值",
            "月度总成本", "最大单笔成本", "最小单笔成本", "备注"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填空单元格标注（预留提示，后续由Office Parser填充）
        fill_font = Font(name="微软雅黑", size=10, color="FF0000")
        fill_cells = [
            (4, 2, "【填空】人力成本均值"), (4, 3, "【填空】物料成本均值"),
            (4, 4, "【填空】运营成本均值"), (4, 5, "【填空】月度总成本"),
            (4, 6, "【填空】最大单笔成本"), (4, 7, "【填空】最小单笔成本"),
            (5, 2, "【填空】人力成本均值"), (5, 3, "【填空】物料成本均值"),
            (5, 4, "【填空】运营成本均值"), (5, 5, "【填空】月度总成本"),
            (5, 6, "【填空】最大单笔成本"), (5, 7, "【填空】最小单笔成本")
        ]
        # 填充部门名称（对应数据库中的部门）
        dept_names = ["销售部", "技术部"]
        for row, dept_name in enumerate(dept_names, 4):
            ws.cell(row=row, column=1, value=dept_name).font = Font(name="微软雅黑", size=10, bold=True)

        # 标记填空单元格
        for row, col, tip in fill_cells:
            cell = ws.cell(row=row, column=col, value=tip)
            cell.font = fill_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 步骤4：调整列宽
        column_widths = [15, 20, 20, 20, 20, 20, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # 步骤5：保存Excel模板
        wb.save(EXCEL_TEMPLATE_PATH)
        logger.info(f"Excel测试模板生成成功，路径：{EXCEL_TEMPLATE_PATH}")

    except Exception as e:
        raise FileOperateException(f"生成Excel模板失败：{str(e)}")

# ---------------------- 主函数：统一执行 ----------------------
def main():
    """主函数：执行数据库初始化 + Excel模板生成"""
    logger = get_logger("Financial_Data_Init_Main")
    try:
        # 步骤1：初始化MySQL数据库
        init_mysql_database()

        # 步骤2：生成Excel测试模板
        generate_excel_test_template()

        logger.info("="*60)
        logger.info("🎉 金融数据初始化全部完成！")
        logger.info(f"✅ 数据库：`{DB_NAME}`（包含3张关联表，批量测试数据）")
        logger.info(f"✅ Excel模板：{EXCEL_TEMPLATE_PATH}（用于填空功能测试）")
        logger.info("💡 后续可直接使用该数据库和Excel模板进行Agent全链路测试")
        logger.info("="*60)

    except Exception as e:
        logger.error(f"❌ 金融数据初始化失败：{str(e)}")

if __name__ == "__main__":
    main()