# ai_report_agent/mcp/office_parser/excel_parser.py（完善，继承BaseMCP）
from typing import Dict, Any, List
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from mcp.base_mcp import BaseMCP
from config import get_office_config
from utils import get_abs_path

class ExcelParser(BaseMCP):
    """Office Parser模块：Excel解析与生成（实现BaseMCP接口）"""
    def __init__(self):
        super().__init__()
        self.template_dir = ""  # 模板目录
        self.supported_formats = ["xlsx", "xls"]  # 支持的Excel格式
        self.workbook = None  # Excel工作簿实例

    def init(self, module_config: Dict[str, Any] = None) -> bool:
        """
        初始化Excel Parser模块（加载模板目录、配置）
        :param module_config: 专属配置（如模板路径）
        :return: 初始化是否成功
        """
        try:
            # 加载Office配置
            self.module_config = get_office_config()
            if module_config:
                self.module_config.update(module_config)
            self.template_dir = get_abs_path(self.module_config.get("excel_template_dir", "./data/templates/excel"))

            # 校验模板目录
            os.makedirs(self.template_dir, exist_ok=True)

            self.initialized = True
            self.logger.info("Office Parser模块（Excel）初始化成功")
            return True
        except Exception as e:
            self.logger.error(f"Office Parser模块（Excel）初始化失败：{str(e)}")
            return False

    def execute(self, task_dict: Dict[str, Any]) -> Dict[str, Any]:
        """
        执行Excel生成任务（标准化输入输出）
        【输入字段】（task_dict["params"]内）：
            - fill_data: 填充数据（分析结果、SQL结果，列表字典/统计字典格式）
            - template_name: 模板文件名（如"cost_analysis_template.xlsx"）
            - sheet_name: 目标工作表名（如"成本分析"）
            - output_filename: 输出文件名（如"2024年2月成本分析.xlsx"）
            - field_mapping: 字段与单元格映射（如{"cost_amount": "B2"}）
        【输出字段】（result["data"]内）：
            - excel_file_path: 生成的Excel文件绝对路径
            - sheet_name: 目标工作表名
            - fill_status: 填充状态（如"全部填充完成"）
            - file_size: 文件大小（单位：KB）
        """
        # 1. 输入标准化
        if not self.initialized:
            return self._standardize_output("failed", error_msg="Excel Parser模块未初始化")
        standardized_task = self._standardize_input(task_dict)
        task_id = standardized_task["task_id"]
        task_params = standardized_task["params"]
        output_path = standardized_task["output_path"]

        try:
            # 2. 提取专属输入参数
            fill_data = task_params.get("fill_data", {})
            template_name = task_params.get("template_name", "default_template.xlsx")
            sheet_name = task_params.get("sheet_name", "Sheet1")
            output_filename = task_params.get("output_filename", f"excel_result_{task_id}.xlsx")
            field_mapping = task_params.get("field_mapping", {})

            # 3. 核心逻辑：加载模板 → 填充数据 → 保存文件
            template_path = os.path.join(self.template_dir, template_name)
            output_full_path = os.path.join(output_path, output_filename)
            os.makedirs(output_path, exist_ok=True)

            # 加载/创建工作簿
            self.workbook = self._load_or_create_workbook(template_path)
            worksheet = self._get_or_create_worksheet(sheet_name)

            # 填充数据（简化实现，支持列表字典和统计字典）
            fill_status = self._fill_excel_data(worksheet, fill_data, field_mapping)

            # 保存Excel文件
            self.workbook.save(output_full_path)
            file_size = round(os.path.getsize(output_full_path) / 1024, 2)  # 转换为KB

            # 4. 构造专属输出数据
            excel_data = {
                "excel_file_path": get_abs_path(output_full_path),
                "sheet_name": sheet_name,
                "fill_status": fill_status,
                "file_size": file_size,
                "output_filename": output_filename
            }

            # 5. 输出标准化
            result = self._standardize_output("success", data=excel_data)
            result["task_id"] = task_id
            result["execute_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            result["output_files"] = [excel_data["excel_file_path"]]

            self.logger.info(f"Excel生成任务 {task_id} 执行成功，文件路径：{excel_data['excel_file_path']}")
            return result
        except Exception as e:
            error_msg = f"Excel生成任务 {task_id} 执行失败：{str(e)}"
            self.logger.error(error_msg)
            result = self._standardize_output("failed", error_msg=error_msg)
            result["task_id"] = task_id
            return result

    def close(self) -> bool:
        """释放Excel资源（关闭工作簿，销毁实例）"""
        try:
            if self.workbook:
                self.workbook.close()
                self.workbook = None
            self.initialized = False
            self.logger.info("Excel Parser模块资源释放成功")
            return True
        except Exception as e:
            self.logger.error(f"Excel Parser模块资源释放失败：{str(e)}")
            return False

    # 以下为辅助方法
    def _load_or_create_workbook(self, template_path: str) -> Workbook:
        """加载模板工作簿，不存在则创建新工作簿"""
        if os.path.exists(template_path) and template_path.endswith(tuple(self.supported_formats)):
            return load_workbook(template_path, data_only=True)
        else:
            self.logger.warning(f"模板 {template_path} 不存在，创建新工作簿")
            return Workbook()

    def _get_or_create_worksheet(self, sheet_name: str):
        """获取工作表，不存在则创建"""
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return self.workbook.create_sheet(title=sheet_name)

    def _fill_excel_data(self, worksheet, fill_data: Any, field_mapping: Dict[str, str]) -> str:
        """填充Excel数据（简化实现）"""
        if isinstance(fill_data, dict):
            # 填充统计字典（如{"cost_amount": {"mean": 15000}}）
            for field, cell in field_mapping.items():
                if field in fill_data:
                    worksheet[cell] = fill_data[field].get("mean", "")
        elif isinstance(fill_data, list):
            # 填充列表字典（如[{"cost_amount": 10000}, ...]）
            for row, data in enumerate(fill_data, start=2):  # 从第2行开始填充
                for col, field in enumerate(field_mapping.keys(), start=1):
                    worksheet.cell(row=row, column=col, value=data.get(field, ""))
        return "全部填充完成"