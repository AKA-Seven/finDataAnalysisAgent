# tools/python_executor.py
"""
进一步修正版 Python执行工具
主要修复：
1. ImportError: __import__ not found – 因为 SAFE_BUILTINS 缺少 __import__，但为了安全，不添加完整 __import__。
   - 解决方案：文档/描述中明确 "无需 import，直接使用预导入模块如 pandas.df(...), plt.plot() 等"。
   - 测试代码已调整为无 import。
2. 增强安全性：保持无 __import__，用户代码不能动态 import 新模块，只用预定义 ALLOWED_MODULES。
3. 输出格式微调：确保 result 包含关键字符串，便于测试匹配。
"""

import sys
import io
import json
import traceback
from pathlib import Path
from typing import Dict
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from utils.db_utils import get_connection, execute_query
from utils.file_utils import get_output_path
from tools.base_tool import BaseTool
from config import config

# 安全内置函数白名单（无 __import__，防止动态导入风险模块）
SAFE_BUILTINS = {
    'print': print,
    'range': range,
    'len': len,
    'str': str,
    'int': int,
    'float': float,
    'list': list,
    'dict': dict,
    'set': set,
    'tuple': tuple,
    'sum': sum,
    'min': min,
    'max': max,
    'abs': abs,
    'round': round,
    'enumerate': enumerate,
    'zip': zip,
    'sorted': sorted,
    'reversed': reversed,
    'any': any,
    'all': all,
    # 添加更多必要内置，如需要
}

# 白名单模块（预导入，用户直接用 pandas 而非 import）
ALLOWED_MODULES = {
    'pandas': pd,
    'plt': plt,
    'load_workbook': load_workbook,
    'get_connection': get_connection,
    'execute_query': execute_query,
    'get_output_path': get_output_path,
    'Path': Path,
    'config': config,
}


class PythonExecutorTool(BaseTool):
    name = "python_executor"
    description = """执行安全的Python数据分析代码。
输入为JSON: {"code": "完整代码字符串"}
无需 import，直接使用预导入模块如 pandas.DataFrame(...), plt.plot(), load_workbook() 等（完整列表见ALLOWED_MODULES）。
代码可处理大数据、字段匹配、Excel补全、绘图。
输出会捕获print内容和生成的文件路径。"""

    def run(self, input_str: str) -> str:
        try:
            data = json.loads(input_str)
            code = data["code"]
        except (json.JSONDecodeError, KeyError):
            return "错误：输入必须为JSON，且包含'code'字段（完整Python代码）"

        # 捕获stdout
        old_stdout = sys.stdout
        sys.stdout = captured_output = io.StringIO()

        try:
            # 受限环境：模块白名单 + 安全builtins（无 __import__）
            exec_globals = ALLOWED_MODULES.copy()
            exec_globals['__builtins__'] = SAFE_BUILTINS

            # 执行代码
            exec(code, exec_globals, {})

            # 保存生成的图表
            fig_paths = []
            for i, fignum in enumerate(plt.get_fignums()):
                fig = plt.figure(fignum)
                fig_path = get_output_path(f"chart_{i + 1}.png")
                fig.savefig(str(fig_path))
                fig_paths.append(str(fig_path))
            plt.close('all')

            # 获取输出
            output = captured_output.getvalue().strip()

            # 构建结果
            result_lines = ["执行成功！"]
            if output:
                result_lines.append("输出内容：")
                result_lines.append(output)
            if fig_paths:
                result_lines.append("生成图表：")
                result_lines.extend(fig_paths)

            return "\n".join(result_lines) if len(result_lines) > 1 else "执行完成，无打印输出。"

        except Exception as e:
            error_detail = traceback.format_exc()
            return f"代码执行错误：{str(e)}\n详细追踪：{error_detail}"
        finally:
            sys.stdout = old_stdout
            plt.close('all')