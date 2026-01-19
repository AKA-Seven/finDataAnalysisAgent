# tools/excel_handler.py
"""
Excel处理工具（支持新建/补全模板，适合大数据写入）
mode: create/update
大数据建议结合python_executor更灵活
"""

import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from utils.file_utils import get_output_path, df_to_table_data
from tools.base_tool import BaseTool

class ExcelHandlerTool(BaseTool):
    name = "excel_handler"
    description = """生成或补全Excel。
输入JSON: {
  "mode": "create"或"update",
  "template_path": "update时必填",
  "sheets": [{"sheet_name": "...", "data": [[...]], "start_cell": "A1"}],
  "output_filename": "data.xlsx"
}"""

    def run(self, input_str: str) -> str:
        try:
            data = json.loads(input_str)
            mode = data.get("mode", "create")
            template_path = data.get("template_path")
            sheets = data.get("sheets", [])
            output_filename = data.get("output_filename", "output.xlsx")
        except json.JSONDecodeError:
            return "错误：输入必须为有效JSON"

        try:
            if mode == "update" and template_path:
                wb = load_workbook(template_path)
            else:
                wb = Workbook()
                wb.remove(wb.active)  # 清空默认sheet

            for sheet_info in sheets:
                sheet_name = sheet_info.get("sheet_name", "Sheet1")
                table_data = sheet_info.get("data", [])
                start_cell = sheet_info.get("start_cell", "A1")

                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]

                # 写入数据
                start_row = int(''.join(filter(str.isdigit, start_cell))) or 1
                start_col = ord(start_cell.upper()[0]) - ord('A') + 1 if start_cell[0].isalpha() else 1

                for r_idx, row in enumerate(table_data):
                    for c_idx, value in enumerate(row):
                        ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

            output_path = get_output_path(output_filename)
            wb.save(str(output_path))
            return f"Excel文件已{'更新' if mode=='update' else '生成'}：{output_path}"
        except Exception as e:
            return f"Excel处理错误：{str(e)}"